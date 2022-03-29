# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import dash
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from dash import dcc, html

app = dash.Dash()

def fig():
    #fig = go.Figure()

    figure = px.scatter(width=1600, height=800)

    figure.add_trace(go.Scatter(x=base_merged_df['tubeid'], y=base_merged_df['intensity1900'],
                                mode='markers', name='Base Intensity 1900 nm', marker=dict(size=16,
                                                                                       line=dict(width=2,
                                                                                                 color='DarkSlateGrey')))
                     )
    figure.add_trace(go.Scatter(x=base_merged_df['tubeid'], y=base_merged_df['intensity1500'],
                                mode='markers', name='Base Intensity 1500 nm', marker=dict(size=8,
                                                                                           line=dict(width=2,
                                                                                                     color='DarkSlateGrey')))
                     )
    figure.add_trace(go.Scatter(x=base_merged_df['tubeid'], y=base_merged_df['intensity1300'],
                                mode='markers', name='Base Intensity 1300 nm', marker=dict(size=8,
                                                                                           line=dict(width=2,
                                                                                                     color='DarkSlateGrey')))
                     )

    figure.add_trace(go.Scatter(x=merged_df['tubeid'], y=merged_df['intensity1300'],
                                mode='markers', name='Intensity 1300 nm', marker=dict(size=8,
                                                                                      line=dict(width=2,
                                                                                                color='DarkSlateGrey')))
                     )
    figure.add_trace(go.Scatter(x=merged_df['tubeid'], y=merged_df['intensity1500'],
                     mode='markers', name='Intensity 1500 nm', marker=dict(size=8,
                                                                           line=dict(width=2,
                                                                                     color='DarkSlateGrey')))
                     )
    figure.add_trace(go.Scatter(x=merged_df['tubeid'], y=merged_df['intensity1900'],
                                mode='markers', name='Intensity 1900 nm', marker=dict(size=16,
                                                                                      line=dict(width=2,
                                                                                                color='DarkSlateGrey')))
                     )

    figure.update_layout(title='''Green region - 1900 nm category (separated by the third quartile),\
     blue - 1300 nm category (first quartile)''',
                         xaxis_title = 'Tube ID',
                         yaxis_title = 'Intensity'
                         )

    figure.add_hrect(y0=minval, y1=q1, line_width=0, fillcolor="blue", opacity=0.2)
    figure.add_hrect(y0=q3, y1=maxval, line_width=0, fillcolor="green", opacity=0.2)

    return figure


def fig_box():
    figurebox = go.Figure()
    figurebox.add_trace(go.Box(y=total_df["intensity1300"], name="1300 nm", boxpoints="all"))
    figurebox.add_trace(go.Box(y=total_df["intensity1500"], name="1500 nm", boxpoints="all"))
    figurebox.add_trace(go.Box(y=total_df["intensity1900"], name="1900 nm * 10", boxpoints="all"))

    return figurebox


if __name__ == '__main__':
    update_excel = False
    run_server = False
    change_only_empty = True
    batch = 2

    base_df1300 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/1_1300sumtot.txt", sep=" ", header=None,
                              names=("tubeid", "intensity1300"))
    base_df1500 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/1_1500sumtot.txt", sep=" ", header=None,
                              names=("tubeid", "intensity1500"))
    base_df1900 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/1_1900sumtot.txt", sep=" ", header=None,
                              names=("tubeid", "intensity1900"))

    base_merged_df = base_df1300.merge(base_df1500)
    base_merged_df = base_merged_df.merge(base_df1900)
    base_merged_df['intensity1900'] = base_merged_df['intensity1900'] * 10
    base_merged_df = base_merged_df.sort_values(by='intensity1900')

    df1300 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/"+str(batch)+"_1300sumtot.txt", sep=" ",
                         header=None, names=("tubeid", "intensity1300"))
    df1500 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/"+str(batch)+"_1500sumtot.txt", sep=" ",
                         header=None, names=("tubeid", "intensity1500"))
    df1900 = pd.read_csv("D:/ADOS-Tech/metrology - Documents/img/analysis/"+str(batch)+"_1900sumtot.txt", sep=" ",
                         header=None, names=("tubeid", "intensity1900"))

    merged_df = df1300.merge(df1500)
    merged_df = merged_df.merge(df1900)
    merged_df['intensity1900'] = merged_df['intensity1900'] * 10
    merged_df = merged_df.sort_values(by='intensity1900')

    total_df = base_merged_df.append(merged_df)
    total_df = total_df.sort_values(by='intensity1900')

    q1 = base_merged_df.quantile(0.25)[2]
    q3 = base_merged_df.quantile(0.75)[2]
    minval = min(base_merged_df['intensity1900'])
    maxval = max(base_merged_df['intensity1900'])

    print(f'q1: {q1}, q3: {q3}, minval: {minval}, maxval: {maxval}')

    list1900 = df1900.loc[df1900['intensity1900'] > q3]
    list1500 = df1900.loc[(df1900['intensity1900'] > q1) & (df1900['intensity1900'] < q3)]
    list1300 = df1900.loc[df1900['intensity1900'] < q1]


    if (update_excel):
        wb = openpyxl.load_workbook('D:/OneDrive - ADOS-Tech/tube_measurements/stock.xlsx')
        ws = wb.active
        end_row = ws.max_row
        start_row = 2
        row_index = start_row

        print(sum(list1300['tubeid'] == "2012_03075"))

        while row_index <= end_row:
            if (ws.cell(row_index,5).value == None):
                if sum(list1300['tubeid'] == ws.cell(row_index, 1).value) > 0:
                    ws.cell(row_index, 5).value = "1300"
                if sum(list1500['tubeid'] == ws.cell(row_index, 1).value) > 0:
                    ws.cell(row_index, 5).value = "1700"
                if sum(list1900['tubeid'] == ws.cell(row_index, 1).value) > 0:
                    ws.cell(row_index, 5).value = "1900"
                print(f'Sno:  {ws.cell(row_index, 1).value},lambda: {ws.cell(row_index, 5).value}')
            row_index += 1
        # TODO: ID with zeroes;
        #df_excel_in['Sno']=df_excel_in['Batch']+"_"+df_excel_in['ID']
        # for row in df_excel_in.itertuples():
        #     #print(row['lambda'])
        #     if not pd.isna(row.Sno):
        #         row.Sno = str(row.Batch)+"_"+str(row.ID)
        #         #print('NO sno')
        #         print(row)

        wb.save('D:/OneDrive - ADOS-Tech/tube_measurements/stock_new.xlsx')
        #df_excel_out = df_excel_in.to_excel("D:/OneDrive - ADOS-Tech/tube_measurements/stock_new.xlsx")
        app.layout = html.Div(id='parent', children=[
            html.H1(id='H1', children='Intensity at different wavelengths', style={'textAlign': 'center',
                                                                                   'marginTop': 40,
                                                                                   'marginBottom': 40}),
            dcc.Graph(id='line_plot1', figure=fig()),
            dcc.Graph(id='box_plot', figure=fig_box())
        ]
                              )
    if run_server:
        app.run_server()
